# -*- coding: utf-8 -*-
"""
Created on Wed Sep 19 16:00:05 2018

@author: lenovo
"""

#from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook 
from openpyxl.styles import NamedStyle
from copy import copy

class SheetStyle:
    
    def  __init__(self,app):
        self.app = app
        mode_sheet = load_workbook('mode.xlsx').active#格式模板
        self.styles = []
        for i in range(5):
            newStyle = NamedStyle(name='Style%d'%(i+1))
            newStyle.fill = copy(mode_sheet['A%d'%(i+1)].fill)
            newStyle.font = copy(mode_sheet['A%d'%(i+1)].font)#字体 
            newStyle.border = copy(mode_sheet['A%d'%(i+1)].border) #边框
            newStyle.alignment = copy(mode_sheet['A%d'%(i+1)].alignment)#对齐
            self.styles.append(newStyle)
    
    '''设置告警文件格式'''    
    def setWarnSheetStyle(self,file_name):
        
        self.app.printInfo('开始设置告警文件表格格式')
        wb = load_workbook(file_name)
        '''加载Style'''
        for style in self.styles:
            wb.add_named_style(style)
        sheet = wb.active
        
        # 设置行高为30
        for inx in range(sheet.max_row):
            sheet.row_dimensions[inx+1].height = 20
        self.app.printInfo('设置行高为30')
        # 设置列宽
        dict_width ={
                'A':15,'B':50,'C':8.5,'D':8.5,
                'E':8.5,'F':13,'G':8.5,'H':13,
                'I':8.5,'J':8.5,'K':8.5,'L':8.5,
                'M':20,'N':15.5,'O':11,'P':15,
                'Q':14.5,'R':11,}
        for k,v in dict_width.items():
            sheet.column_dimensions[k].width = v
        self.app.printInfo('设置每列宽度')
        #设置每行格式
        for row in sheet.iter_rows():
            cell_style = self.styles[2]
            
            if row[-1].value == 4:
                cell_style = self.styles[3]
            elif row[-1].value == 5:
                cell_style = self.styles[4]
            elif row[-1].value == 1:
                cell_style = self.styles[0]
            elif row[-1].value == 2:
                cell_style = self.styles[1]
            for cell in row:
                cell.style = cell_style
        self.app.printInfo('设置每行格式')
        # 删除最后一列
        sheet.max_column
        sheet.delete_cols(sheet.max_column)
        wb.save(file_name)
        self.app.printInfo('告警文件格式处理完成')
        self.app.printInfo('已保存入' + file_name)
        return file_name
        
        